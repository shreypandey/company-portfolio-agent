"""build_excel_workbook — multi-sheet Excel from a structured payload.

Schema is flexible: the LLM passes whatever sections it has data for, missing ones
are skipped gracefully. Supports up to 10 sheets for US public companies.
"""
from __future__ import annotations

import asyncio
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..config import OUTPUTS_DIR
from ..context import RunContext
from ..envelope import Result

NAME = "build_excel_workbook"

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_LABEL_FONT = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=14, color="1F4E78")


def _write_header(ws, row: int, cols: list[str]) -> None:
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws, min_w: int = 10, max_w: int = 60) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)


def _coerce(v: Any) -> Any:
    """Convert to something Excel can store. Numbers stay numbers; complex -> str."""
    if v is None:
        return None
    if isinstance(v, (int, float, str, bool)):
        return v
    return str(v)


# ──────── Sheet writers ────────

def _is_empty_sheet(ws) -> bool:
    """A sheet is 'empty' if it has only the header row or nothing."""
    if ws.max_row <= 1:
        return True
    # Check if rows 2..N have any non-null data
    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 50), values_only=True):
        if any(v not in (None, "") for v in row):
            return False
    return True


def _write_overview(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Overview")
    ws.cell(row=1, column=1, value=data.get("company") or data.get("name") or "Company").font = _TITLE_FONT
    _write_header(ws, 3, ["Field", "Value"])
    rows: list[tuple] = [("Company", data.get("company") or data.get("name"))]
    for k, v in (data.get("overview") or {}).items():
        rows.append((k, v))
    for i, (k, v) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=str(k)).font = _LABEL_FONT
        ws.cell(row=i, column=2, value=_coerce(v))
    _autosize(ws)


def _write_statement_sheet(wb: Workbook, sheet_name: str, statement: dict | None) -> None:
    """Statement format: {periods: [...], rows: [{label, values}, ...], currency?}"""
    statement = _unwrap(statement) if statement else None
    if not statement or not statement.get("periods"):
        return
    ws = wb.create_sheet(sheet_name)
    periods = statement["periods"]
    cols = ["Line Item", *periods]
    _write_header(ws, 1, cols)
    rows = statement.get("rows") or []
    for r, row in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=row.get("label", "")).font = _LABEL_FONT
        for c, val in enumerate(row.get("values") or [], start=2):
            ws.cell(row=r, column=c, value=_coerce(val))
    # currency footer
    if statement.get("currency"):
        ws.cell(row=len(rows) + 3, column=1, value=f"Currency: {statement['currency']}").font = Font(italic=True)
    _autosize(ws)


def _write_ratios_sheet(wb: Workbook, sheet_name: str, ratios: dict | None) -> None:
    """Accept either statement shape ({periods, rows}) or simple rows-only data."""
    ratios = _unwrap(ratios) if ratios else None
    if not ratios:
        return
    if ratios.get("periods"):
        _write_statement_sheet(wb, sheet_name, ratios)
        return
    rows = ratios.get("rows") or []
    if not rows:
        return
    ws = wb.create_sheet(sheet_name)
    _write_header(ws, 1, ["Metric", "Value"])
    for r, row in enumerate(rows, start=2):
        label = row.get("label", "")
        values = row.get("values") or []
        value = None
        for v in values:
            if v not in (None, ""):
                value = v
                break
        ws.cell(row=r, column=1, value=label).font = _LABEL_FONT
        ws.cell(row=r, column=2, value=_coerce(value))
    _autosize(ws)


def _write_earnings_sheet(wb: Workbook, earnings: dict | None) -> None:
    earnings = _unwrap(earnings) if earnings else None
    if not earnings or not earnings.get("quarters"):
        return
    ws = wb.create_sheet("Earnings Track")
    cols = ["Period", "EPS Actual", "EPS Estimated", "Surprise", "Surprise %", "Beat?"]
    _write_header(ws, 1, cols)
    for r, q in enumerate(earnings["quarters"], start=2):
        # Accept both "period" (Finnhub/our tool) and "date" (FMP-ish)
        ws.cell(row=r, column=1, value=_coerce(q.get("period") or q.get("date")))
        ws.cell(row=r, column=2, value=_coerce(q.get("eps_actual")))
        ws.cell(row=r, column=3, value=_coerce(q.get("eps_estimated")))
        ws.cell(row=r, column=4, value=_coerce(q.get("surprise")))
        ws.cell(row=r, column=5, value=_coerce(q.get("surprise_pct")))
        ws.cell(row=r, column=6, value="✓" if q.get("beat") else "✗")
    _autosize(ws)


def _unwrap(d: dict | None) -> dict:
    """The LLM sometimes wraps tool output, e.g. {ok, data: {...}}. Peel one level if present."""
    if not d:
        return {}
    if isinstance(d, dict) and "data" in d and isinstance(d["data"], dict) and ("ok" in d or "sources_used" in d):
        return d["data"]
    return d


def _write_analyst_sheet(wb: Workbook, analyst: dict | None) -> None:
    analyst = _unwrap(analyst)
    if not analyst:
        return
    ws = wb.create_sheet("Analyst Coverage")
    row = 1

    # Latest distribution summary (Finnhub: latest_total_analysts + latest_distribution_pct)
    total = analyst.get("latest_total_analysts")
    dist_pct = analyst.get("latest_distribution_pct") or {}
    if total or dist_pct:
        ws.cell(row=row, column=1, value="Latest Consensus Summary").font = _TITLE_FONT
        row += 1
        _write_header(ws, row, ["Metric", "Value"])
        row += 1
        if total:
            ws.cell(row=row, column=1, value="Total Analysts").font = _LABEL_FONT
            ws.cell(row=row, column=2, value=_coerce(total))
            row += 1
        if analyst.get("latest_period"):
            ws.cell(row=row, column=1, value="Latest Period").font = _LABEL_FONT
            ws.cell(row=row, column=2, value=_coerce(analyst["latest_period"]))
            row += 1
        for k in ("strongBuy", "buy", "hold", "sell", "strongSell"):
            if k in dist_pct:
                ws.cell(row=row, column=1, value=f"{k} (%)").font = _LABEL_FONT
                ws.cell(row=row, column=2, value=_coerce(dist_pct[k]))
                row += 1
        row += 1

    # Price target summary (FMP shape — kept for paid-tier upgrade path)
    pt = analyst.get("price_target") or {}
    if pt:
        ws.cell(row=row, column=1, value="Price Target Summary").font = _TITLE_FONT
        row += 1
        _write_header(ws, row, ["Metric", "Value"])
        row += 1
        for k in ("count", "mean", "high", "low"):
            if k in pt:
                ws.cell(row=row, column=1, value=k).font = _LABEL_FONT
                ws.cell(row=row, column=2, value=_coerce(pt[k]))
                row += 1
        row += 1

    # Monthly recommendations — accept both Finnhub keys (period/strongBuy/buy/...) and FMP keys.
    recs = analyst.get("monthly_recommendations") or []
    if recs:
        ws.cell(row=row, column=1, value="Monthly Recommendations").font = _TITLE_FONT
        row += 1
        cols = ["Period", "Strong Buy", "Buy", "Hold", "Sell", "Strong Sell"]
        _write_header(ws, row, cols)
        row += 1
        for r in recs:
            # date / period (Finnhub: "period", FMP: "date")
            period = r.get("period") or r.get("date")
            sb = r.get("strongBuy", r.get("analystRatingsStrongBuy"))
            b = r.get("buy", r.get("analystRatingsBuy"))
            h = r.get("hold", r.get("analystRatingsHold"))
            s = r.get("sell", r.get("analystRatingsSell"))
            ss = r.get("strongSell", r.get("analystRatingsStrongSell"))
            ws.cell(row=row, column=1, value=_coerce(period))
            ws.cell(row=row, column=2, value=_coerce(sb))
            ws.cell(row=row, column=3, value=_coerce(b))
            ws.cell(row=row, column=4, value=_coerce(h))
            ws.cell(row=row, column=5, value=_coerce(s))
            ws.cell(row=row, column=6, value=_coerce(ss))
            row += 1
        row += 2

    # Notes/limitations from the tool (e.g. "price targets unavailable on free tier")
    note = analyst.get("note")
    if note:
        ws.cell(row=row, column=1, value="Note").font = _LABEL_FONT
        ws.cell(row=row, column=2, value=str(note))
        row += 1

    # Fallback for agent-supplied "quarters" payloads that look more like a combined
    # earnings / analyst snapshot than the canonical monthly_recommendations shape.
    quarters = analyst.get("quarters") or []
    if quarters and not (total or dist_pct or pt or recs):
        ws.cell(row=row, column=1, value="Analyst / Earnings Snapshot").font = _TITLE_FONT
        row += 1
        _write_header(ws, row, ["Date", "Consensus Rating", "Surprise %", "Source"])
        row += 1
        for q in quarters:
            ws.cell(row=row, column=1, value=_coerce(q.get("date") or q.get("period")))
            ws.cell(row=row, column=2, value=_coerce(q.get("consensus_rating")))
            ws.cell(row=row, column=3, value=_coerce(q.get("surprise_pct")))
            ws.cell(row=row, column=4, value=_coerce(q.get("source")))
            row += 1
        row += 1

    _autosize(ws)


def _collect_source_rows(node: Any, *, path: str = "", rows: list[tuple] | None = None) -> list[tuple]:
    """Collect traceable source-ish records from the payload."""
    if rows is None:
        rows = []
    if isinstance(node, dict):
        source_name = node.get("source") or node.get("source_name") or node.get("sourceType")
        source_url = node.get("url") or node.get("source_url") or node.get("website")
        if source_name or source_url:
            tool_used = node.get("tool") or node.get("tool_used") or path.split(".", 1)[0] or "payload"
            data_extracted = ", ".join(
                str(k) for k in ("name", "ticker", "headline", "title", "date", "period") if node.get(k) not in (None, "")
            ) or path or "payload"
            rows.append(
                (
                    source_name or "derived",
                    source_url or "",
                    node.get("source_type") or ("company website" if "website" in node else "payload"),
                    tool_used,
                    data_extracted,
                    "",
                    "",
                )
            )
        for k, v in node.items():
            if k in {"source", "source_name", "source_url", "sourceType"}:
                continue
            _collect_source_rows(v, path=f"{path}.{k}" if path else k, rows=rows)
    elif isinstance(node, list):
        for i, item in enumerate(node):
            _collect_source_rows(item, path=f"{path}[{i}]" if path else str(i), rows=rows)
    return rows


def _write_source_log_sheet(wb: Workbook, payload: dict) -> None:
    ws = wb.create_sheet("Source Log")
    _write_header(
        ws,
        1,
        ["Source name", "Source URL", "Source type", "Tool used", "Data extracted", "Access date", "Reliability notes"],
    )
    rows = _collect_source_rows(payload)
    if not rows:
        rows = [("derived", "", "payload", "build_excel_workbook", "no source rows exposed in payload", "", "")]
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=_coerce(val))
    _autosize(ws)


def _write_data_gaps_sheet(wb: Workbook, payload: dict, populated_sheets: set[str]) -> None:
    ws = wb.create_sheet("Data Gaps")
    _write_header(ws, 1, ["Missing data", "Why it matters", "Reason unavailable", "Best available substitute", "Confidence level"])

    gaps: list[tuple[str, str, str, str, str]] = []
    if not payload.get("income_statement_quarterly"):
        gaps.append(("Quarterly financials", "Needed to inspect near-term revenue and margin changes", "No quarterly series was passed into the workbook tool", "Annual statements only", "Medium"))
    if not payload.get("key_ratios"):
        gaps.append(("Key ratios history", "Needed for historical margins, leverage, and efficiency trends", "The workbook payload did not include a complete ratio series", "Annual statements and overview metrics", "Medium"))
    if not payload.get("analyst"):
        gaps.append(("Analyst coverage", "Needed for consensus and rating trend context", "No analyst payload was supplied", "Financial and earnings data", "Low"))
    if not payload.get("ownership"):
        gaps.append(("Ownership and leadership", "Needed for insider / holder context", "No ownership payload was supplied", "Company overview", "Low"))
    if not payload.get("macro"):
        gaps.append(("Sector and macro context", "Needed to contextualize the company against the broader environment", "No macro payload was supplied", "Company and sector narrative", "Low"))
    if "Recent Developments" not in populated_sheets:
        gaps.append(("Recent developments", "Needed for latest events and catalysts", "The agent payload did not include a dedicated recent-development section", "Source log and portfolio narrative", "Medium"))
    if "Source Log" not in populated_sheets:
        gaps.append(("Source log", "Needed for traceability", "Source metadata was only partially exposed in the payload", "Payload-derived source rows", "Medium"))

    if not gaps:
        gaps.append(("None", "All modeled sections were populated", "N/A", "N/A", "High"))

    for r, gap in enumerate(gaps, start=2):
        for c, val in enumerate(gap, start=1):
            ws.cell(row=r, column=c, value=_coerce(val))
    _autosize(ws)


def _write_ownership_sheet(wb: Workbook, ownership: dict | None) -> None:
    ownership = _unwrap(ownership)
    if not ownership:
        return
    ws = wb.create_sheet("Ownership & Insider")
    row = 1

    # Insider summary (our tool returns this)
    summary = ownership.get("insider_summary") or {}
    if summary:
        ws.cell(row=row, column=1, value="Insider Activity Summary").font = _TITLE_FONT
        row += 1
        _write_header(ws, row, ["Metric", "Value"])
        row += 1
        for k, v in summary.items():
            ws.cell(row=row, column=1, value=str(k)).font = _LABEL_FONT
            ws.cell(row=row, column=2, value=_coerce(v))
            row += 1
        row += 1

    # Recent insider transactions — Finnhub keys: name, transactionDate, filingDate,
    # transactionCode (P=purchase, S=sale), transactionPrice, change (shares ±), share (post-tx)
    insider = ownership.get("insider_recent") or []
    if insider:
        ws.cell(row=row, column=1, value="Recent Insider Transactions").font = _TITLE_FONT
        row += 1
        _write_header(ws, row, ["Transaction Date", "Filed", "Insider", "Type", "Shares", "Price", "Value (USD)"])
        row += 1
        for r in insider[:30]:
            # Support both Finnhub and FMP-ish shapes
            name = r.get("name") or r.get("reportingName")
            tx_date = r.get("transactionDate") or r.get("date")
            filed = r.get("filingDate")
            ttype = r.get("transactionCode") or r.get("transactionType")
            shares = r.get("change") if r.get("change") is not None else r.get("securitiesTransacted")
            price = r.get("transactionPrice") if r.get("transactionPrice") is not None else r.get("price")
            ws.cell(row=row, column=1, value=_coerce(tx_date))
            ws.cell(row=row, column=2, value=_coerce(filed))
            ws.cell(row=row, column=3, value=_coerce(name))
            ws.cell(row=row, column=4, value=_coerce(ttype))
            ws.cell(row=row, column=5, value=_coerce(shares))
            ws.cell(row=row, column=6, value=_coerce(price))
            try:
                value = float(shares) * float(price) if shares not in (None, 0) and price not in (None, 0) else None
            except (TypeError, ValueError):
                value = None
            ws.cell(row=row, column=7, value=_coerce(value))
            row += 1
        row += 2

    # Fallback for agent-supplied simplified ownership payloads.
    major = ownership.get("major_insiders") or []
    if major:
        ws.cell(row=row, column=1, value="Major Insiders / Leadership").font = _TITLE_FONT
        row += 1
        _write_header(ws, row, ["Name", "Type", "Role", "Ownership / Shares", "Recent Activity", "Notes"])
        row += 1
        for item in major:
            name = item.get("name")
            role = item.get("role")
            typ = "Insider" if role else "Holder"
            shares = item.get("shares") or item.get("ownership_percent")
            recent = item.get("recent_activity") or ownership.get("insider_sentiment") or ""
            note = item.get("notes") or ""
            ws.cell(row=row, column=1, value=_coerce(name))
            ws.cell(row=row, column=2, value=_coerce(typ))
            ws.cell(row=row, column=3, value=_coerce(role))
            ws.cell(row=row, column=4, value=_coerce(shares))
            ws.cell(row=row, column=5, value=_coerce(recent))
            ws.cell(row=row, column=6, value=_coerce(note))
            row += 1
        row += 1

    # Monthly insider sentiment (Finnhub)
    sentiment = ownership.get("insider_sentiment_monthly") or []
    if sentiment:
        ws.cell(row=row, column=1, value="Monthly Insider Sentiment").font = _TITLE_FONT
        row += 1
        _write_header(ws, row, ["Year-Month", "Net Change", "MSPR"])
        row += 1
        for s in sentiment:
            ym = f"{s.get('year')}-{str(s.get('month','')).zfill(2)}"
            ws.cell(row=row, column=1, value=ym)
            ws.cell(row=row, column=2, value=_coerce(s.get("change")))
            ws.cell(row=row, column=3, value=_coerce(s.get("mspr")))
            row += 1
        row += 2

    # Top institutional holders (paid endpoint — included for forward compatibility)
    inst = ownership.get("institutional_top_holders") or []
    if inst:
        ws.cell(row=row, column=1, value="Top Institutional Holders").font = _TITLE_FONT
        row += 1
        _write_header(ws, row, ["Holder", "Shares", "Date Reported", "Change"])
        row += 1
        for h in inst[:30]:
            ws.cell(row=row, column=1, value=_coerce(h.get("holder")))
            ws.cell(row=row, column=2, value=_coerce(h.get("shares")))
            ws.cell(row=row, column=3, value=_coerce(h.get("dateReported")))
            ws.cell(row=row, column=4, value=_coerce(h.get("change")))
            row += 1
        row += 2

    note = ownership.get("note")
    if note:
        ws.cell(row=row, column=1, value="Note").font = _LABEL_FONT
        ws.cell(row=row, column=2, value=str(note))
    _autosize(ws)


def _write_peers_sheet(wb: Workbook, competitors: dict | None) -> None:
    competitors = _unwrap(competitors) if competitors else None
    if not competitors:
        return
    peers = competitors.get("peers_enriched") or competitors.get("peers") or []
    if not peers:
        return
    ws = wb.create_sheet("Peers Comparison")
    # Collect all keys across peers (union)
    cols: list[str] = []
    seen: set[str] = set()
    for p in peers:
        for k in p.keys():
            if k not in seen:
                seen.add(k)
                cols.append(k)
    _write_header(ws, 1, cols)
    for r, p in enumerate(peers, start=2):
        for c, k in enumerate(cols, start=1):
            ws.cell(row=r, column=c, value=_coerce(p.get(k)))
    _autosize(ws)


def _write_macro_sheet(wb: Workbook, macro: dict | None) -> None:
    macro = _unwrap(macro) if macro else None
    if not macro or not macro.get("series"):
        return
    ws = wb.create_sheet("Macro Context")
    ws.cell(row=1, column=1, value=f"Sector: {macro.get('sector_resolved') or macro.get('sector_requested')}").font = _TITLE_FONT
    row = 3
    for series in macro["series"]:
        ws.cell(row=row, column=1, value=f"{series.get('series_id')} — {series.get('label', '')}").font = _LABEL_FONT
        if series.get("why_relevant"):
            ws.cell(row=row, column=2, value=f"why: {series['why_relevant']}").font = Font(italic=True)
        row += 1
        _write_header(ws, row, ["Date", "Value"])
        row += 1
        for o in (series.get("observations") or []):
            ws.cell(row=row, column=1, value=_coerce(o.get("date")))
            ws.cell(row=row, column=2, value=_coerce(o.get("value")))
            row += 1
        row += 1
    _autosize(ws)


# ──────── Tool entry ────────

SCHEMA = {
    "name": NAME,
    "description": (
        "Build a multi-sheet Excel workbook from the structured data you've gathered. Sheets you can "
        "include: Overview, Income Statement (annual + quarterly), Balance Sheet (annual + quarterly), "
        "Cash Flow (annual + quarterly), Key Ratios History, Earnings Track, Analyst Coverage, "
        "Ownership & Insider, Peers Comparison, Macro Context. Pass only the sections you have data for; "
        "missing ones are skipped gracefully. Call this once near the end of research."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "data": {
                "type": "object",
                "properties": {
                    "company": {"type": "string"},
                    "overview": {"type": "object", "description": "Flat key/value: ticker, country, industry, employees, market_cap, etc.", "additionalProperties": True},
                    "income_statement_annual": {"type": "object", "description": "{periods, rows: [{label, values}]}", "additionalProperties": True},
                    "income_statement_quarterly": {"type": "object", "additionalProperties": True},
                    "balance_sheet_annual": {"type": "object", "additionalProperties": True},
                    "balance_sheet_quarterly": {"type": "object", "additionalProperties": True},
                    "cash_flow_annual": {"type": "object", "additionalProperties": True},
                    "cash_flow_quarterly": {"type": "object", "additionalProperties": True},
                    "key_ratios": {"type": "object", "additionalProperties": True},
                    "earnings": {"type": "object", "description": "{quarters: [{date, eps_actual, eps_estimated, surprise, surprise_pct, beat}]}", "additionalProperties": True},
                    "analyst": {"type": "object", "additionalProperties": True},
                    "ownership": {"type": "object", "additionalProperties": True},
                    "competitors": {"type": "object", "description": "{peers_enriched: [{ticker, name, market_cap, pe, ...}]}", "additionalProperties": True},
                    "macro": {"type": "object", "additionalProperties": True},
                },
                "required": ["company"],
            }
        },
        "required": ["data"],
    },
}


def _build(payload: dict, out_path: Path) -> dict:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet

    _write_overview(wb, payload)
    _write_statement_sheet(wb, "Income Statement (A)", payload.get("income_statement_annual"))
    _write_statement_sheet(wb, "Income Statement (Q)", payload.get("income_statement_quarterly"))
    _write_statement_sheet(wb, "Balance Sheet (A)", payload.get("balance_sheet_annual"))
    _write_statement_sheet(wb, "Balance Sheet (Q)", payload.get("balance_sheet_quarterly"))
    _write_statement_sheet(wb, "Cash Flow (A)", payload.get("cash_flow_annual"))
    _write_statement_sheet(wb, "Cash Flow (Q)", payload.get("cash_flow_quarterly"))
    _write_ratios_sheet(wb, "Key Ratios History", payload.get("key_ratios"))
    _write_earnings_sheet(wb, payload.get("earnings"))
    _write_analyst_sheet(wb, payload.get("analyst"))
    _write_ownership_sheet(wb, payload.get("ownership"))
    _write_peers_sheet(wb, payload.get("competitors"))
    _write_macro_sheet(wb, payload.get("macro"))
    _write_source_log_sheet(wb, payload)

    if not wb.sheetnames:
        ws = wb.create_sheet("Overview")
        ws["A1"] = "No data available for any sheet."

    # Detect which sheets ended up empty (header only / no data rows)
    empty: list[str] = []
    for name in wb.sheetnames:
        if _is_empty_sheet(wb[name]):
            empty.append(name)

    # Data gaps should always be present and populated.
    _write_data_gaps_sheet(wb, payload, set(wb.sheetnames) - set(empty))

    wb.save(out_path)
    return {
        "path": str(out_path),
        "sheets": wb.sheetnames,
        "empty_sheets": empty,
        "populated_sheets": [s for s in wb.sheetnames if s not in empty],
        "size_bytes": out_path.stat().st_size,
    }


async def run(ctx: RunContext, data: dict) -> dict:
    run_dir = OUTPUTS_DIR / ctx.run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    safe = "".join(c for c in (data.get("company") or "company") if c.isalnum() or c in " _-")[:60].strip().replace(" ", "_") or "company"
    out_path = run_dir / f"{safe}_financials.xlsx"
    try:
        meta = await asyncio.to_thread(_build, data, out_path)
    except Exception as e:
        return Result.failure(
            sources_attempted=["openpyxl"],
            errors_by_source={"openpyxl": f"{type(e).__name__}: {e}"},
            hint="Excel generation failed. Inspect the structure of `data`.",
        ).to_dict()

    return Result.success(meta, sources_used=["openpyxl"]).to_dict()
