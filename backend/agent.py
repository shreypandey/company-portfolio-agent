"""Agent loop: provider-agnostic tool-use ReAct with streaming + per-turn checkpoint.

The agent talks to whichever LLM provider is selected via `LLM_PROVIDER` (anthropic
or gemini). Provider-specific format conversion lives in `backend/llm/*`.

Events emitted via the sink (forwarded to the SSE layer):

  started          — run_id, company, model, provider
  iteration        — n
  agent_text       — text delta (assistant prose)
  tool_start       — tool block began streaming (id, name)
  tool_input_delta — partial JSON of tool input
  tool_input_done  — tool input block complete
  tool_executing   — about to run tool (id, name, input)
  source_try       — internal: tool is trying a source (tool_id, source)
  source_ok        — internal: source succeeded
  source_miss      — internal: source had no data (clean miss)
  source_error     — internal: source errored
  source_cache_hit — internal: served from per-run cache
  tool_result      — tool finished (id, name, preview, is_error)
  turn_done        — assistant turn finished (stop_reason, usage)
  retry            — LLM API retry happening
  error            — non-fatal error
  done             — final paths
"""
from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Awaitable, Callable

import httpx

from .config import (
    MAX_ITERATIONS,
    MAX_OUTPUT_TOKENS,
    OUTPUTS_DIR,
    RUNS_DIR,
    SSE_PREVIEW_CHARS,
)
from .context import RunContext
from .llm import get_llm_client
from .tools import registry

log = logging.getLogger("agent.loop")

EventSink = Callable[[str, dict], Awaitable[None]]

PROMPT_PATH = Path(__file__).with_name("prompt.md")


def build_system_prompt() -> str:
    """Load the file-backed prompt and attach run-time context."""
    prompt_template = PROMPT_PATH.read_text(encoding="utf-8").strip()
    current_timestamp = datetime.now().astimezone().isoformat(timespec="seconds")
    return f"{prompt_template}\n\nRUNTIME CONTEXT\nCurrent timestamp: {current_timestamp}"



class Agent:
    # When the model's input_tokens cross this fraction of its context window,
    # archive older tool results to a 1-line summary to free room.
    CONTEXT_TRIM_THRESHOLD = 0.80
    # Keep this many most-recent (assistant, tool_result) pairs intact when trimming.
    CONTEXT_PRESERVE_PAIRS = 4

    def __init__(self, company: str, run_id: str, sink: EventSink):
        self.company = company
        self.run_id = run_id
        self.sink = sink
        self.llm = get_llm_client()
        self.system_prompt = build_system_prompt()
        self.run_dir = RUNS_DIR / run_id
        self.run_dir.mkdir(parents=True, exist_ok=True)
        self.out_dir = OUTPUTS_DIR / run_id
        self.out_dir.mkdir(parents=True, exist_ok=True)
        self.messages: list[dict] = []
        self.fatal_error: str | None = None
        self.http: httpx.AsyncClient | None = None
        self.ctx: RunContext | None = None
        # tool_use_id → original full result. Mirrored to runs/{run_id}/archive.json
        # whenever we archive older tool_results to summaries.
        self.archive: dict[str, dict] = {}

    def _checkpoint(self) -> None:
        try:
            (self.run_dir / "state.json").write_text(
                json.dumps(
                    {
                        "run_id": self.run_id,
                        "company": self.company,
                        "provider": self.llm.provider,
                        "model": self.llm.model,
                        "messages": self.messages,
                    },
                    indent=2,
                    default=str,
                )
            )
        except Exception:
            log.exception("checkpoint write failed")

    async def _execute_tool_uses(self, content_blocks: list[dict]) -> list[dict]:
        assert self.ctx is not None
        results: list[dict] = []
        for b in content_blocks:
            if b.get("type") != "tool_use":
                continue
            tool_input = b.get("input") or {}
            tool_id = b["id"]
            tool_name = b["name"]
            await self.sink("tool_executing", {"id": tool_id, "name": tool_name, "input": tool_input})

            self.ctx.parent_tool_id = tool_id
            try:
                result = await registry.dispatch(tool_name, tool_input, self.ctx)
            finally:
                self.ctx.parent_tool_id = None

            content_str = json.dumps(result, default=str)
            is_error = isinstance(result, dict) and (
                result.get("ok") is False or bool(result.get("error"))
            )
            await self.sink(
                "tool_result",
                {
                    "id": tool_id,
                    "name": tool_name,
                    "is_error": is_error,
                    "preview": content_str[:SSE_PREVIEW_CHARS],
                    "size": len(content_str),
                },
            )
            results.append(
                {
                    "tool_use_id": tool_id,
                    "tool_name": tool_name,
                    "content": content_str,
                    "is_error": is_error,
                }
            )
        return results

    def _summarize_archived(self, tool_name: str, content_str: str) -> str:
        """Build a 1-line summary used in place of an archived tool result.

        Cheap deterministic — peeks at ok/sources/keys without spending an LLM call.
        """
        size_kb = max(1, len(content_str) // 1024)
        try:
            obj = json.loads(content_str)
        except Exception:
            return f"[archived] tool={tool_name} ~{size_kb}KB (see runs/{self.run_id}/archive.json)"
        if not isinstance(obj, dict):
            return f"[archived] tool={tool_name} ~{size_kb}KB"
        ok = obj.get("ok")
        sources = obj.get("sources_used") or []
        data = obj.get("data") if isinstance(obj.get("data"), dict) else {}
        keys = list(data.keys())[:6] if data else []
        hint = obj.get("hint")
        parts = [f"[archived] tool={tool_name}", f"~{size_kb}KB"]
        if ok is not None:
            parts.append(f"ok={ok}")
        if sources:
            parts.append(f"sources={sources}")
        if keys:
            parts.append(f"keys={keys}")
        if not ok and hint:
            parts.append(f"hint={hint[:80]}")
        parts.append(f"see runs/{self.run_id}/archive.json")
        return " | ".join(parts)

    def _maybe_archive_old_tool_results(self, last_usage: dict | None) -> None:
        """If input_tokens > THRESHOLD * window, replace older tool_result content
        with a 1-line summary. Original content moves to self.archive (and disk).

        Triggered after each turn_done. Idempotent: tool_results already marked
        '[archived]' are skipped.
        """
        if not last_usage:
            return
        input_tokens = last_usage.get("input_tokens") or 0
        window = self.llm.context_window or 0
        if not window or input_tokens < window * self.CONTEXT_TRIM_THRESHOLD:
            return

        # Locate (assistant, user-with-tool-results) pairs in order.
        pair_indices: list[tuple[int, int]] = []
        i = 0
        while i < len(self.messages) - 1:
            a = self.messages[i]
            u = self.messages[i + 1]
            if a.get("role") == "assistant" and u.get("role") == "user" and isinstance(u.get("content"), list):
                if any(isinstance(b, dict) and b.get("type") == "tool_result" for b in u["content"]):
                    pair_indices.append((i, i + 1))
                    i += 2
                    continue
            i += 1

        if len(pair_indices) <= self.CONTEXT_PRESERVE_PAIRS:
            return  # nothing old enough to trim

        to_trim = pair_indices[: -self.CONTEXT_PRESERVE_PAIRS]
        trimmed = 0
        bytes_freed = 0
        for assist_idx, user_idx in to_trim:
            name_by_id: dict[str, str] = {}
            for b in self.messages[assist_idx].get("content", []):
                if isinstance(b, dict) and b.get("type") == "tool_use":
                    name_by_id[b.get("id", "")] = b.get("name", "")
            for block in self.messages[user_idx]["content"]:
                if not (isinstance(block, dict) and block.get("type") == "tool_result"):
                    continue
                original = block.get("content")
                if not isinstance(original, str) or original.startswith("[archived]"):
                    continue
                tool_use_id = block.get("tool_use_id", "")
                tool_name = name_by_id.get(tool_use_id, "unknown_tool")
                self.archive[tool_use_id] = {
                    "tool_name": tool_name,
                    "content": original,
                    "size_bytes": len(original),
                }
                bytes_freed += len(original)
                block["content"] = self._summarize_archived(tool_name, original)
                trimmed += 1

        if trimmed == 0:
            return

        # Persist the archive to disk
        try:
            (self.run_dir / "archive.json").write_text(
                json.dumps(self.archive, indent=2, default=str)
            )
        except Exception:
            log.exception("archive write failed")

        return {
            "trimmed_count": trimmed,
            "bytes_freed_approx": bytes_freed,
            "input_tokens": input_tokens,
            "window": window,
            "pct_used_before": round(input_tokens / window * 100, 1),
        }

    async def run(self) -> dict:
        async with httpx.AsyncClient(timeout=30.0, http2=False) as http:
            self.http = http
            self.ctx = RunContext(run_id=self.run_id, sink=self.sink, http=http)

            self.messages.append(
                {"role": "user", "content": f"Research this company and produce the deliverables: {self.company}"}
            )
            await self.sink(
                "started",
                {
                    "run_id": self.run_id,
                    "company": self.company,
                    "model": self.llm.model,
                    "provider": self.llm.provider,
                },
            )

            final_text_parts: list[str] = []
            last_stop: str | None = None
            iteration = 0

            for iteration in range(1, MAX_ITERATIONS + 1):
                await self.sink("iteration", {"n": iteration, "max": MAX_ITERATIONS})

                try:
                    final = await self.llm.stream_turn(
                        system=self.system_prompt,
                        messages=self.messages,
                        tools=registry.SCHEMAS,
                        max_tokens=MAX_OUTPUT_TOKENS,
                        sink=self.sink,
                    )
                except Exception as e:
                    log.exception("LLM call failed terminally")
                    self.fatal_error = f"{type(e).__name__}: {e}"
                    await self.sink("error", {"fatal": True, "message": self.fatal_error})
                    break

                # capture the latest turn's text (final portfolio)
                turn_text_blocks = [b["text"] for b in final.content_blocks if b.get("type") == "text"]
                if turn_text_blocks:
                    final_text_parts = turn_text_blocks

                self.messages.append({"role": "assistant", "content": final.content_blocks})

                await self.sink("turn_done", {"stop_reason": final.stop_reason, "usage": final.usage})
                last_stop = final.stop_reason

                # If we're closing on the context window, archive older tool_results
                # to summaries. Keeps the run alive on long agentic loops.
                trim_report = self._maybe_archive_old_tool_results(final.usage)
                if trim_report:
                    await self.sink("context_trim", trim_report)

                if final.stop_reason == "tool_use":
                    results = await self._execute_tool_uses(final.content_blocks)
                    # Provider-specific encoding of tool results into next message(s)
                    self.messages.extend(self.llm.encode_tool_results(results))
                    self._checkpoint()
                    continue

                break
            else:
                await self.sink("error", {"fatal": False, "message": f"hit max iterations ({MAX_ITERATIONS})"})

        # Write portfolio — strip any leading preamble before the first markdown heading
        portfolio_md = "\n\n".join(final_text_parts).strip()
        if portfolio_md:
            lines = portfolio_md.split("\n")
            for i, line in enumerate(lines):
                if line.lstrip().startswith("#"):
                    portfolio_md = "\n".join(lines[i:]).strip()
                    break
        portfolio_md = portfolio_md or f"# {self.company}\n\n_No content generated._"
        portfolio_path = self.out_dir / "portfolio.md"
        portfolio_path.write_text(portfolio_md)
        self._checkpoint()

        excel_files = sorted(self.out_dir.glob("*.xlsx"))
        excel_path = str(excel_files[0]) if excel_files else None

        # Post-inspect the workbook to surface empty sheets.
        excel_sheets: list[str] = []
        excel_empty_sheets: list[str] = []
        if excel_files:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(excel_files[0], read_only=True)
                for name in wb.sheetnames:
                    ws = wb[name]
                    excel_sheets.append(name)
                    is_empty = True
                    if ws.max_row > 1:
                        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 50), values_only=True):
                            if any(v not in (None, "") for v in row):
                                is_empty = False
                                break
                    if is_empty:
                        excel_empty_sheets.append(name)
            except Exception:
                log.exception("post-run sheet inspection failed")

        await self.sink(
            "done",
            {
                "run_id": self.run_id,
                "portfolio_path": str(portfolio_path),
                "excel_path": excel_path,
                "excel_sheets": excel_sheets,
                "excel_empty_sheets": excel_empty_sheets,
                "iterations": iteration,
                "last_stop_reason": last_stop,
                "aborted": bool(self.fatal_error),
                "error": self.fatal_error,
                "provider": self.llm.provider,
                "model": self.llm.model,
            },
        )
        return {
            "portfolio_path": str(portfolio_path),
            "excel_path": excel_path,
            "iterations": iteration,
        }
